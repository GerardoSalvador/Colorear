#Fecha de creación: 09/02/2023
#José Gerardo Salvador Pérez Sánchez
#Villahermosa, Tabasco

#Creado para colorear las celdas de un excel cuando el identificador
#cambie.

#Librerias necesarias para esta solución
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import os
os.system('cls' if os.name == 'nt' else 'clear')


def colorear():
    ruta = input("Dame la ruta del archivo excel (.xlsx):\n")
    libro = load_workbook(ruta)
    hoja = libro.active


    #se inicia en la fila 2 puesto que la fila 1 es de encabezados
    contador_filas = 2
    valor_fila = (hoja[f'A{contador_filas}'].value)
    i=0


    #lista donde se guardaran todos los identificadores de los usuarios
    identificadores = []
    i
    #while que se detiene hasta que encuentre una celda vacia
    while(valor_fila != None):
        valor_fila = (hoja[f'A{contador_filas}'].value)
        if (valor_fila == None):
            break
        else:
            valor_fila = int(hoja[f'A{contador_filas}'].value)
            identificadores.append(valor_fila)
            contador_filas += 1
            i += 1


    #Lista que almacenará los valores de la columna id sin repeticiones
    identificadores_no_repetido = []
    #for que evaluará si algun valor esta repetido en la primera lista y solo se guardará una vez en la lista nueva
    for i in identificadores:
        if i not in identificadores_no_repetido:
            identificadores_no_repetido.append(i)
    #variable que contiene la cantidad total de identificadores sin repetir
    numero_total_identificadores_no_repetidos = len(identificadores_no_repetido)


    #Lista que almacenará valores que esten en una posición par de la lista identificadores_no_repetidos
    identificadores_posicion_par = []
    #for que evaluará y agregará a la lista identificadores_posicion_par los valores que esten en una posicion divisible entre 2
    for i in range(0, numero_total_identificadores_no_repetidos):
        if i % 2 == 0:
            valor = identificadores_no_repetido[i]
            identificadores_posicion_par.insert(i, valor)
        else:
            continue


    for i in range(2, contador_filas):
        id = int(hoja[f'A{i}'].value)
        if id not in identificadores_posicion_par:
            hoja[f'A{i}'].fill = PatternFill(patternType='solid', fgColor='9EAEF4')
            hoja[f'B{i}'].fill = PatternFill(patternType='solid', fgColor='9EAEF4')
            hoja[f'C{i}'].fill = PatternFill(patternType='solid', fgColor='9EAEF4')
            hoja[f'D{i}'].fill = PatternFill(patternType='solid', fgColor='9EAEF4')
            hoja[f'E{i}'].fill = PatternFill(patternType='solid', fgColor='9EAEF4')


    #La ruta dada al principio del script es donde se guardará el archivo coloreado
    #ademas se sobreescribirá
    libro.save(ruta)


total=int(input("Cuantos archivos excel deseas colorear:"))
for i in range(1,total+1):
    print("\nExcel #",i)
    colorear()
    print("Excel #",i,"ha sido sobreescrito correctamente")
